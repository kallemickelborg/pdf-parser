"""XLSX exporter for validated canonical invoices.

Columns match the shape requested by the project spec:
    Invoice_no | Vendor | Customer | Due_Date | Gross_Total_Amount | Billing_Type | Currency
"""

from __future__ import annotations

import io
from collections.abc import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.domain import ParseResult

EXPORT_COLUMNS: tuple[str, ...] = (
    "Invoice_no",
    "Vendor",
    "Customer",
    "Due_Date",
    "Gross_Total_Amount",
    "Billing_Type",
    "Currency",
)


def _row(result: ParseResult) -> list[str]:
    inv = result.invoice
    return [
        inv.invoice_no or "",
        inv.vendor or "",
        inv.customer or "",
        inv.due_date.isoformat() if inv.due_date else "",
        f"{inv.gross_total_amount:.2f}" if inv.gross_total_amount is not None else "",
        inv.billing_type or "",
        inv.currency or "",
    ]


def export_to_xlsx(
    results: Iterable[ParseResult],
    *,
    include_partial: bool = False,
) -> bytes:
    """Build an XLSX workbook for the given parse results and return its bytes.

    Only `parsed` results are included by default (fully validated canonical
    fields). Pass `include_partial=True` to also include partials.
    """
    allowed_statuses = {"parsed"} | ({"partial"} if include_partial else set())

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    ws.append(list(EXPORT_COLUMNS))
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for result in results:
        if result.status in allowed_statuses:
            ws.append(_row(result))

    for i, column in enumerate(EXPORT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(len(column) + 2, 18)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
